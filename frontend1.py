import pandas as pd
from openpyxl import load_workbook
import streamlit as st
import requests
import json
import os
import pandas as pd
import numpy as np
import google.generativeai as genai
from sklearn.metrics.pairwise import cosine_similarity
import mysql.connector
from mysql.connector import Error
import streamlit as st
import pandas as pd
import io
import numpy as np
from datetime import datetime
import streamlit as st
from openpyxl import load_workbook, Workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from io import BytesIO
import tempfile
from copy import copy
from datetime import datetime
import streamlit as st
import pandas as pd
import io
import openpyxl
from openpyxl.styles import Font
import re
import traceback
from copy import copy
import streamlit as st
import os
import google.generativeai as genai
from PyPDF2 import PdfReader
from dotenv import load_dotenv
import validators,streamlit as st
from langchain.prompts import PromptTemplate
from langchain_groq import ChatGroq
from langchain.chains.summarize import load_summarize_chain
from langchain_community.document_loaders import YoutubeLoader,UnstructuredURLLoader
import re
import time
import datetime
from db import get_db_connection
import datetime
from typing import List, Optional

load_dotenv()
GEMINI_API_KEY = os.getenv("GEMINI_API_KEY")
genai.configure(api_key=GEMINI_API_KEY)

API_URL = "http://localhost:5000"

st.set_page_config(page_title="Knowledge Sharing Platform", layout="centered")

# --- Session State Initialization ---
if 'page_history' not in st.session_state:
    st.session_state.page_history = []
if "token" not in st.session_state:
    st.session_state.token = None
if "show_signup" not in st.session_state:
    st.session_state.show_signup = False
if "user_email" not in st.session_state:
    st.session_state.user_email = None
if "user_id" not in st.session_state:
    st.session_state.user_id = None

st.title("...🐛DataMorph ...")

# --- Helper Functions ---
# Simple Gemini Wrapper class without model specification

db_connection = get_db_connection()

def param_gen():
        def generate_m_query(parameter_name, sql_column_name):
                var_safe_param_name = parameter_name.replace(" ", "_")
                param_ref = parameter_name
                
                code = f"""
            {var_safe_param_name}_list = if Type.Is(Value.Type({param_ref}), List.Type) then
                {param_ref}
            else
                {{{param_ref}}},
                
            isSelectAll_{var_safe_param_name} = List.Contains({var_safe_param_name}_list, "__SelectAll__"),

            where_clause_{var_safe_param_name} = if isSelectAll_{var_safe_param_name}
            then " 1 = 1 "
            else
                let
                    CleanedList = List.Transform({var_safe_param_name}_list, each Text.Replace(_, "'", "")),
                    
                    AddSingleQuotes = List.Transform(
                        CleanedList, each "'" & _ & "'"),
                        
                    DelimitedList = Text.Combine(AddSingleQuotes, ","),

                    formatted = " {sql_column_name} IN ( " &
                        DelimitedList
                    & " ) "
                in
                    formatted,
            """
                return code

        def process_sql_query(sql_query, parameter_details):
                processed_query = sql_query
                
                for param_name, column_name, sql_var_name in parameter_details:
                    var_safe_param_name = re.sub(r'\s+', '_', param_name)
                    column_name_escaped = re.escape(column_name)
                    
                    pattern = f"{column_name_escaped}\\s+(?i:in)\\s*\\(\"&\\s*{sql_var_name}\\s*&\"\\)"
                    replacement = f"\" & where_clause_{var_safe_param_name} & \""
                    
                    processed_query = re.sub(pattern, replacement, processed_query, flags=re.IGNORECASE)
                
                return processed_query

        def remove_sql_comments(sql_query):
                lines = sql_query.split('\n')
                filtered_lines = [line for line in lines if not line.strip().startswith('--')]
                return '\n'.join(filtered_lines)

        def validate_parameter_name(name):
                if " " in name:
                    return False
                return True

            # --- Streamlit UI ---
        

            # Add Help Button
        with st.expander("📚 How to Use (Help)", expanded=False):
                st.markdown("""
                ### Application Guide
                1. **Add Parameters**: Click the 'Add Parameter' button for each parameter you want to use.
                2. **Fill Details**:
                - **Parameter Name**: Must not have spaces.
                - **SQL Column Name**: Column name in the SQL query.
                - **SQL Variable Name**: Variable name placeholder used inside the SQL query.
                3. **Enter SQL Query**: Paste your base SQL query (optional).
                4. **Options**:
                - Check 'Remove comments' if you want to strip lines starting with `--`.
                5. **Click 'Generate M Query'** to generate:
                - Complete M Query code
                - Individual M Query snippets
                - Updated/processed SQL Query.

                ### Behind the Scenes
                - For each parameter:
                - A **where_clause** is created.
                - If `__SelectAll__` is selected → returns `1 = 1`.
                - Else → generates a `column_name IN ('val1','val2')` clause.
                - In SQL query:
                - The code looks for patterns like `ColumnName IN ("& sql_var &")` and replaces them dynamically with `where_clause_paramName`.
                """)

        st.write("Generate M query code based on multiple parameters and SQL column names.")
        st.write("Modify the SQL query to include the generated M query code.")

            # Initialize session state
        if 'parameters' not in st.session_state:
                st.session_state.parameters = []

            # Add parameter button
        if st.button("Add Parameter"):
                st.session_state.parameters.append({
                    "param_name": f"parameter_name_{len(st.session_state.parameters)+1}", 
                    "column_name": f"COLUMN-NAME-{len(st.session_state.parameters)+1}",
                    "sql_var_name": f"sql_var_{len(st.session_state.parameters)+1}"
                })

            # Parameters input
        st.subheader("Parameters")
        st.info("Note: Parameter names cannot contain spaces.")

        parameter_error = False

        for i, param in enumerate(st.session_state.parameters):
                col1, col2, col3 = st.columns(3)
                with col1:
                    param_name = st.text_input(f"Parameter Name {i+1}:", param["param_name"], key=f"param_{i}")
                    if not validate_parameter_name(param_name):
                        st.error("Parameter name cannot contain spaces.")
                        parameter_error = True
                with col2:
                    column_name = st.text_input(f"SQL Column Name {i+1}:", param["column_name"], key=f"column_{i}")
                with col3:
                    sql_var_name = st.text_input(f"SQL Variable Name {i+1}:", param.get("sql_var_name", param["param_name"]), key=f"sql_var_{i}")
                
                st.session_state.parameters[i]["param_name"] = param_name
                st.session_state.parameters[i]["column_name"] = column_name
                st.session_state.parameters[i]["sql_var_name"] = sql_var_name

            # SQL query input
        st.subheader("SQL Query (Optional)")
        default_query = """"""

        sql_query = st.text_area("Enter your SQL query:", default_query, height=150)

        remove_comments = st.checkbox("Remove comments from SQL query", value=False, 
                                        help="When checked, removes all lines that start with '--' from the SQL query")

        generate_button = st.button("Generate M Query")

        if generate_button and not parameter_error:
                individual_m_queries = {}
                all_code = ""
                processed_query = ""
                
                for param in st.session_state.parameters:
                    param_code = generate_m_query(param["param_name"], param["column_name"])
                    individual_m_queries[param["param_name"]] = param_code
                    all_code += param_code + "\n"
                
                if sql_query:
                    working_sql_query = sql_query
                    if remove_comments:
                        working_sql_query = remove_sql_comments(sql_query)
                    
                    param_details = [(p["param_name"], p["column_name"], p["sql_var_name"]) for p in st.session_state.parameters]
                    processed_query = process_sql_query(working_sql_query, param_details)
                
                tab1, tab2, tab3 = st.tabs(["Complete M Query", "Individual M Queries", "Processed SQL Query"])
                
                with tab2:
                    if individual_m_queries:
                        for param_name, code in individual_m_queries.items():
                            with st.expander(f"M Query for {param_name}"):
                                st.code(code, language="m")
                    else:
                        st.info("No parameters added yet. Add parameters to generate individual M queries.")
                
                with tab1:
                    if all_code:
                        st.code(all_code, language="m")
                    else:
                        st.info("No parameters added yet. Add parameters to generate M query code.")
                
                with tab3:
                    if processed_query:
                        if remove_comments and sql_query != working_sql_query:
                            st.subheader("Original SQL Query")
                            st.code(sql_query, language="sql")
                            
                            st.subheader("Query with Comments Removed")
                            st.code(working_sql_query, language="sql")
                            
                            st.subheader("Final Processed SQL Query")
                            st.code(processed_query, language="sql")
                        else:
                            st.code(processed_query, language="sql")
                    else:
                        st.info("Either no SQL query provided or no parameters added to process the SQL query.")
        elif generate_button and parameter_error:
                st.error("Please fix the parameter name errors before generating M Query code.")
def YT_and_WebSummarization_through_Link():
            groq_api_key = st.text_input("Groq API Key",value="",type="password")

            llm = ChatGroq(model="gemma2-9b-it",groq_api_key=groq_api_key)
            st.subheader("summarize URL")
            generic_url  = st.text_input("URL",label_visibility="collapsed")

            prompt_template = """
            Provide summary of the following content in 300 words:
            Content:{text}
            """
            prompt = PromptTemplate(template=prompt_template,input_variables=["text"])

            if st.button("Summarize the content from YT or Website"):
                ##validate all the inputs 
                if not groq_api_key.strip() or not generic_url.strip():
                    st.error("please provide the information to get started")
                elif not validators.url(generic_url):
                    st.error("please enter a valid url.It can maybe a YT video URL or website URL")
                else:
                    try:
                        with st.spinner("Waiting. . ."):
                            ##loading the website or yt video data
                            if "youtube.com" in generic_url or "youtu.be" in generic_url:
                                loader = YoutubeLoader.from_youtube_url(generic_url, add_video_info=True)
                            else:
                                loader = UnstructuredURLLoader(
                                    urls=[generic_url],
                                    ssl_verified=True,
                                    headers={"User-Agent": "Mozilla/5.0"}
                                )
                            docs = loader.load()
                            ## chain for summarization 
                            chain = load_summarize_chain(llm,chain_type="stuff",prompt = prompt)
                            output_summary = chain.run(docs)
                            st.success(output_summary)
                    except Exception as e:
                        st.exception(f"Exception:{e}") 
def Excel_Combiner():
            st.write("This app allows you to upload multiple Excel workbooks, rename the sheets, and combine them into one.")

            # --- Helper Function for Sanitizing Sheet Names ---
            def sanitize_sheet_name(name):
                """Removes invalid characters and truncates sheet names for Excel."""
                # Remove invalid characters: \ / * ? : [ ]
                name = re.sub(r'[\\/*?:\[\]]', '_', name)
                # Truncate to Excel's 31-character limit
                return name[:31]

            # --- Helper Function to Copy Sheet with Images ---
            def copy_sheet_with_images(source_sheet, target_workbook, new_sheet_name):
                """Copy a worksheet including images, charts and all cell styles."""
                # Create a new sheet in the target workbook
                target_sheet = target_workbook.create_sheet(title=new_sheet_name)
                
                # Copy cell values, styles, dimensions
                for row in source_sheet.rows:
                    for cell in row:
                        new_cell = target_sheet.cell(row=cell.row, column=cell.column, value=cell.value)
                        if cell.has_style:
                            new_cell.font = copy(cell.font)
                            new_cell.border = copy(cell.border)
                            new_cell.fill = copy(cell.fill)
                            new_cell.number_format = copy(cell.number_format)
                            new_cell.protection = copy(cell.protection)
                            new_cell.alignment = copy(cell.alignment)
                
                # Copy column dimensions
                for key, dimension in source_sheet.column_dimensions.items():
                    target_sheet.column_dimensions[key].width = dimension.width
                    target_sheet.column_dimensions[key].hidden = dimension.hidden
                
                # Copy row dimensions
                for key, dimension in source_sheet.row_dimensions.items():
                    target_sheet.row_dimensions[key].height = dimension.height
                    target_sheet.row_dimensions[key].hidden = dimension.hidden
                
                # Copy merged cells
                for merged_cell_range in source_sheet.merged_cells.ranges:
                    target_sheet.merge_cells(str(merged_cell_range))
                
                # Copy images and charts
                if source_sheet._images:
                    for image in source_sheet._images:
                        target_sheet.add_image(copy(image))
                
                # Handle other drawing objects
                if source_sheet._charts:
                    for chart in source_sheet._charts:
                        target_sheet.add_chart(copy(chart))
                
                return target_sheet

            # --- Streamlit UI ---
            # Add file naming inputs in a horizontal layout
            col1, col2 = st.columns(2)
                
            with col1:
                model_name = st.text_input("Model Name", "")
                
            with col2:
                report_name = st.text_input("Report Name", "")

            num_pages = st.number_input("How many Excel workbooks do you want to process?", min_value=1, value=2, step=1)

            uploaded_files = []
            page_names = []

            cols = st.columns(num_pages)

            for i in range(num_pages):
                with cols[i]:
                    st.subheader(f"Workbook {i+1}")
                    # Use a more descriptive default page name if desired
                    default_page_name = f"Source_{i+1}"
                    page_name = st.text_input(f"Suffix for sheets from Workbook {i+1}", value=default_page_name, key=f"page_name_{i}")
                    uploaded_file = st.file_uploader(f"Upload Excel workbook {i+1}", type=["xlsx", "xls"], key=f"file_{i}")

                    uploaded_files.append(uploaded_file)
                    page_names.append(page_name)

            # --- Processing Logic ---
            if st.button("Process and Combine Workbooks"):
                # Check if all files are uploaded
                if None in uploaded_files:
                    st.error("Please upload all required Excel workbooks.")
                else:
                    progress_bar = st.progress(0)
                    status_text = st.empty()
                    processed_sheets_count = 0
                    total_sheets_estimate = 0 # Estimate total sheets for progress bar

                    try:
                        # --- Step 1: Estimate total sheets for progress bar ---
                        status_text.text("Analyzing input files...")
                        temp_total_sheets = 0
                        for i, uploaded_file in enumerate(uploaded_files):
                            # Need to reset file pointer after reading names
                            uploaded_file.seek(0)
                            try:
                                xls_temp = pd.ExcelFile(uploaded_file)
                                temp_total_sheets += len(xls_temp.sheet_names)
                            except Exception as e:
                                st.warning(f"Could not read sheet names from Workbook {i+1}. Skipping estimation for this file. Error: {e}")
                            finally:
                                uploaded_file.seek(0) # IMPORTANT: Reset pointer again for actual processing
                        total_sheets_estimate = temp_total_sheets if temp_total_sheets > 0 else 1 # Avoid division by zero

                        # --- Step 2: Create the combined workbook ---
                        status_text.text("Creating combined workbook structure...")
                        combined_wb = openpyxl.Workbook()
                        # Remove the default sheet created by openpyxl
                        if "Sheet" in combined_wb.sheetnames:
                            default_sheet = combined_wb["Sheet"]
                            combined_wb.remove(default_sheet)

                        # Keep track of sheet names used in the *new* workbook to avoid duplicates
                        final_sheet_names_in_workbook = set()
                        all_sheet_names_added = [] # List to show the user which sheets were added

                        # --- Step 3: Process each uploaded workbook ---
                        for i, (uploaded_file, page_name) in enumerate(zip(uploaded_files, page_names)):
                            status_text.text(f"Processing Workbook {i+1} ('{page_name}')...")
                            uploaded_file.seek(0) # Ensure file pointer is at the beginning

                            try:
                                # Save the uploaded file to a temporary location and use openpyxl to open it
                                temp_data = uploaded_file.read()
                                with io.BytesIO(temp_data) as temp_file:
                                    # Load workbook with openpyxl
                                    source_wb = openpyxl.load_workbook(temp_file, data_only=False)
                                    sheet_names = source_wb.sheetnames
                                    
                                    # Process each sheet in the current workbook
                                    for sheet_index, sheet_name in enumerate(sheet_names):
                                        current_sheet_progress = (processed_sheets_count + 1) / total_sheets_estimate
                                        progress_bar.progress(min(current_sheet_progress, 1.0)) # Cap progress at 1.0
                                        status_text.text(f"Processing Workbook {i+1} ('{page_name}') - Sheet: '{sheet_name}'...")
                                        
                                        source_sheet = source_wb[sheet_name]
                                        
                                        # Skip empty sheets
                                        if source_sheet.max_row <= 1 and source_sheet.max_column <= 1:
                                            # Check if only cell A1 exists and is empty
                                            if source_sheet.max_row == 1 and source_sheet.max_column == 1:
                                                if source_sheet.cell(row=1, column=1).value is None and not source_sheet._images:
                                                    st.info(f"Skipping empty sheet: '{sheet_name}' from Workbook {i+1}")
                                                    processed_sheets_count += 1
                                                    continue
                                            else:
                                                st.info(f"Skipping empty sheet: '{sheet_name}' from Workbook {i+1}")
                                                processed_sheets_count += 1
                                                continue

                                        # --- Create and sanitize the new sheet name ---
                                        base_new_sheet_name = f"{sheet_name}_{page_name}"
                                        sanitized_base_name = sanitize_sheet_name(base_new_sheet_name)

                                        # Ensure uniqueness *after* sanitization/truncation
                                        final_sheet_name = sanitized_base_name
                                        counter = 1
                                        while final_sheet_name in final_sheet_names_in_workbook:
                                            suffix = f"_{counter}"
                                            # Ensure the base name + suffix doesn't exceed 31 chars
                                            truncate_at = 31 - len(suffix)
                                            if truncate_at <= 0:
                                                # Handle edge case where suffix itself makes it too long (should be rare)
                                                final_sheet_name = f"Sheet_{processed_sheets_count+1}"[:31] # Fallback name
                                            else:
                                                final_sheet_name = sanitized_base_name[:truncate_at] + suffix
                                            counter += 1
                                            if counter > 100: # Safety break to prevent infinite loops
                                                st.warning(f"Could not generate unique name for sheet derived from '{sheet_name}'/'{page_name}'. Using fallback.")
                                                final_sheet_name = f"Sheet_{processed_sheets_count+1}"[:31]
                                                while final_sheet_name in final_sheet_names_in_workbook:
                                                    processed_sheets_count +=1 # Just ensure uniqueness
                                                    final_sheet_name = f"Sheet_{processed_sheets_count+1}"[:31]
                                                break # Exit inner while loop

                                        # --- Copy sheet to combined workbook with images ---
                                        copy_sheet_with_images(source_sheet, combined_wb, final_sheet_name)
                                        final_sheet_names_in_workbook.add(final_sheet_name)
                                        all_sheet_names_added.append(final_sheet_name)
                                        processed_sheets_count += 1

                            except Exception as sheet_error:
                                st.warning(f"Could not process Workbook {i+1} ('{page_name}'). Error: {sheet_error}. Skipping this workbook.")
                                st.code(traceback.format_exc())
                                continue # Move to the next workbook

                        progress_bar.progress(1.0) # Ensure progress bar reaches 100%
                        status_text.text("Combining complete. Preparing download...")

                        # --- Step 4: Save and provide download ---
                        if not all_sheet_names_added:
                            st.warning("No data sheets were found or processed in the uploaded workbooks.")
                            status_text.text("") # Clear status
                            progress_bar.empty() # Remove progress bar

                        else:
                            with io.BytesIO() as output:
                                combined_wb.save(output)
                                output.seek(0)
                                data = output.getvalue()

                            st.success("All processable workbooks and sheets combined successfully!")
                            st.write(f"Combined workbook contains the following sheets: {', '.join(sorted(all_sheet_names_added))}")

                            # Generate dynamic filename with current date
                            current_date = datetime.now().strftime("%d-%m-%Y")
                            
                            # Clean model and report names (remove spaces or special characters)
                            clean_model_name = re.sub(r'[\\/*?:"<>|\s]', ' ', model_name) if model_name else "Model"
                            clean_report_name = re.sub(r'[\\/*?:"<>|\s]', ' ', report_name) if report_name else "Report"
                            
                            dynamic_filename = f"{clean_model_name}-{clean_report_name}-{current_date}.xlsx"

                            st.download_button(
                                label="Download Combined Workbook",
                                data=data,
                                file_name=dynamic_filename,
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )
                            status_text.text("") # Clear status text on success
                            progress_bar.empty() # Remove progress bar

                    except Exception as e:
                        st.error(f"An unexpected error occurred during processing: {e}")
                        st.error("Please check your input files. If the problem persists, check the logs or report the error.")
                        st.code(traceback.format_exc()) # Show detailed error traceback for debugging
                        status_text.text("Processing failed.")
                        progress_bar.empty() # Remove progress bar
def Generate_Validation_Report():
            header_text = """
        <p class="big-font">
            📝 Important Assumptions:
        </p>
        <ol>
            <li> 📂 Upload two independent Excel files: "Cognos Data" and "PBI Data".</li>
            <li> 🏷️ Make sure the column names are similar in both sheets.</li>
            <li> 📌 Select columns to build the ID part.</li>
            <li>If the Title and Sub-title in Cognos are different, replicate the Sub-title with font size 16 and bold and color - black.</li>
            <li>Replicate the "$" symbol if it is present in the Cognos report.</li>
            <li>If the "$" symbol is present and the value is negative (negative currency numbers), we are using parentheses by default. However, if negative non-currency numbers in Cognos are shown in parentheses, we simply display the default negative numbers without replicating the parentheses.</li>
            <li>Matrix validation should be done in Excel: Sheet 1: Checklist, Sheet 2: Screenshots of both Power BI and Cognos.</li>
            <li>Please create a sheet and add Screenshots of PBI and Cognos in every Validation report.</li>
        </ol>
        """
            header_text = header_text.replace("<li>If the Title and Sub-title in Cognos are different, replicate the Sub-title with font size 16 and bold and color - black.</li>", "<li>🎨 If the Title and Sub-title in Cognos are different, replicate the Sub-title with font size 16 and bold and color - black.</li>")
            header_text = header_text.replace("<li>Replicate the \"$\" symbol if it is present in the Cognos report.</li>", "<li>💲 Replicate the \"$\" symbol if it is present in the Cognos report.</li>")
            header_text = header_text.replace("<li>If the \"$\" symbol is present and the value is negative (negative currency numbers), we are using parentheses by default. However, if negative non-currency numbers in Cognos are shown in parentheses, we simply display the default negative numbers without replicating the parentheses.</li>", "<li>📉 If the \"$\" symbol is present and the value is negative (negative currency numbers), we are using parentheses by default. However, if negative non-currency numbers in Cognos are shown in parentheses, we simply display the default negative numbers without replicating the parentheses.</li>")
            header_text = header_text.replace("<li>Matrix validation should be done in Excel: Sheet 1: Checklist, Sheet 2: Screenshots of both Power BI and Cognos.</li>", "<li>📊 Matrix validation should be done in Excel: Sheet 1: Checklist, Sheet 2: Screenshots of both Power BI and Cognos.</li>")
            header_text = header_text.replace("<li>Please create a sheet and add Screenshots of PBI and Cognos in every Validation report.</li>", "<li> 📸 Please create a sheet and add Screenshots of PBI and Cognos in every Validation report.</li>")

            st.markdown(header_text, unsafe_allow_html=True)

            model_name = st.text_input("Enter the model name:")
            report_name = st.text_input("Enter the report name:")
            
            
            cognos_file = st.file_uploader("Upload Cognos Data Excel file 📈", type="xlsx", key="cognos_upload")
            pbi_file = st.file_uploader("Upload PBI Data Excel file 📉", type="xlsx", key="pbi_upload")
            dry_run(cognos_file, pbi_file)

            if cognos_file is not None and pbi_file is not None:
                try:
                    cognos_df = pd.read_excel(cognos_file)
                    pbi_df = pd.read_excel(pbi_file)

                    # Convert numeric-like strings
                    cognos_df = convert_possible_numeric(cognos_df)
                    pbi_df = convert_possible_numeric(pbi_df)

                    # Standardize text columns
                    cognos_df = cognos_df.apply(lambda x: x.str.upper().str.strip() if x.dtype == "object" else x)
                    pbi_df = pbi_df.apply(lambda x: x.str.upper().str.strip() if x.dtype == "object" else x)

                    option_data = st.radio("Select Option 🛠️", ["Data Present 📊", "Only Column Names Present 🏷️"])

                    if option_data == "Only Column Names Present 🏷️":
                        column_checklist_df = column_checklist(cognos_df, pbi_df)

                        st.subheader("Column Checklist Preview 📋")
                        st.dataframe(column_checklist_df)

                        output = io.BytesIO()
                        with pd.ExcelWriter(output, engine='openpyxl') as writer:
                            checklist_df.to_excel(writer, sheet_name='Checklist', index=False)
                            cognos_df.to_excel(writer, sheet_name='Cognos', index=False)
                            pbi_df.to_excel(writer, sheet_name='PBI', index=False)
                            column_checklist_df.to_excel(writer, sheet_name='Column Checklist', index=False)
                            pd.DataFrame().to_excel(writer, sheet_name='Cognos SS', index=False)
                            pd.DataFrame().to_excel(writer, sheet_name='PBI SS', index=False)

                        output.seek(0)
                        today_date = datetime.date.today().strftime('%Y-%m-%d')
                        dynamic_filename = f"{model_name}_{report_name}_ColumnCheck_Report_{today_date}.xlsx" if model_name and report_name else f"ColumnCheck_Report_{today_date}.xlsx"

                        st.download_button(
                            label="Download Column Check Excel Report ⬇️",
                            data=output,
                            file_name=dynamic_filename,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                        st.info("Once downloaded, please browse to your downloads folder to access the report.")

                    elif option_data == "Data Present 📊":
                        common_columns = list(set(cognos_df.columns) & set(pbi_df.columns))
                        selected_columns = st.multiselect("Select columns to build the ID part: 🔑", common_columns)
                        if selected_columns:
                            validation_report, cognos_agg, pbi_agg = generate_validation_report(cognos_df, pbi_df, selected_columns)
                            column_checklist_df = column_checklist(cognos_df, pbi_df)
                            diff_checker_df = generate_diff_checker(validation_report)

                            st.subheader("Validation Report Preview 📈📉")
                            st.dataframe(validation_report)

                            # Checklist input section
                            # st.subheader("Checklist Status 📝")
                            # for i, row in checklist_df.iterrows():
                            #     checklist_df.loc[i, 'Status - Level1'] = st.text_input(f"Status - Level1 for {row['Checklist']} ✅", key=f"level1_{i}")
                            #     checklist_df.loc[i, 'Status - Level2'] = "" #Only Level 1 input.

                            output = io.BytesIO()
                            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                                checklist_df.to_excel(writer, sheet_name='Checklist', index=False)
                                cognos_agg.to_excel(writer, sheet_name='Cognos', index=False)
                                pbi_agg.to_excel(writer, sheet_name='PBI', index=False)
                                validation_report.to_excel(writer, sheet_name='Validation_Report', index=False)
                                column_checklist_df.to_excel(writer, sheet_name='Column Checklist', index=False)
                                diff_checker_df.to_excel(writer, sheet_name='Diff Checker', index=False)
                                pd.DataFrame().to_excel(writer, sheet_name='Cognos SS', index=False)
                                pd.DataFrame().to_excel(writer, sheet_name='PBI SS', index=False)

                            output.seek(0)
                            today_date = datetime.date.today().strftime('%Y-%m-%d')
                            dynamic_filename = f"{model_name}_{report_name}_ValidationReport_{today_date}.xlsx" if model_name and report_name else f"ValidationReport_{today_date}.xlsx"

                            st.download_button(
                                label="Download Excel Report ⬇️",
                                data=output,
                                file_name=dynamic_filename,
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )
                            st.info("Once downloaded, please browse to your downloads folder to access the report.")

                        else:
                            st.warning("Please select at least one column to build the ID part. ⚠️")
                    
                    

                except Exception as e:
                    st.error(f"An error occurred: {str(e)} ❌")
        
                st.header("Validation Report Analyzer")
                uploaded_file = st.file_uploader("Upload a CSV or Excel file", type=["csv", "xlsx"])

                if uploaded_file is not None:
                    file_content = uploaded_file.getvalue()
                    file_type = uploaded_file.name.split('.')[-1].lower()

                    # Analyze the data
                    results = analyze_validation_report(file_content, file_type)

                    # Print the results
                    if results['error']:
                        st.error(f"Error: {results['error']}")
                    else:
                        st.header("Analysis Results:")
                        st.dataframe(results['comparison_df'])

                        st.subheader("Presence Analysis")
                        st.write(f"Number of 'Present in PBI' entries: {results['presence_analysis']['pbi_present_count']}")
                        st.write(f"Number of 'Present in Cognos' entries: {results['presence_analysis']['cognos_present_count']}")

                        if results['presence_analysis']['pbi_present_ids']:
                            st.write("Unique IDs 'Present in PBI':")
                            st.write(results['presence_analysis']['pbi_present_ids'])
                        else:
                            st.write("No Unique IDs 'Present in PBI'.")

                        if results['presence_analysis']['cognos_present_ids']:
                            st.write("Unique IDs 'Present in Cognos':")
                            st.write(results['presence_analysis']['cognos_present_ids'])
                            st.title("🧩 Unique Key Segment Comparator")

                                    # Input fields
                            cognos_key = st.text_input("Enter Cognos Unique Key")
                            pbi_key = st.text_input("Enter PBI Unique Key")

                            def extract_mismatched_segment(cognos_key, pbi_key):
                                cognos_parts = cognos_key.split('-')
                                pbi_parts = pbi_key.split('-')

                                mismatch_info = []

                                min_len = min(len(cognos_parts), len(pbi_parts))

                                for i in range(min_len):
                                    if cognos_parts[i] != pbi_parts[i]:
                                        column_name = selected_columns[i] if i < len(selected_columns) else f'Segment {i+1}'
                                        mismatch_info.append({
                                            'segment_position': i + 1,
                                            'cognos_value': cognos_parts[i],
                                            'pbi_value': pbi_parts[i],
                                            'length_mismatch': len(cognos_parts[i]) != len(pbi_parts[i]),
                                            'column_name': column_name  # Get the column name from selected_columns
                                        })
                                
                                # Check if one key has more segments than the other
                                if len(cognos_parts) > len(pbi_parts):
                                    for i in range(min_len, len(cognos_parts)):
                                        column_name = selected_columns[i] if i < len(selected_columns) else f'Segment {i+1}'
                                        mismatch_info.append({
                                            'segment_position': i + 1,
                                            'cognos_value': cognos_parts[i],
                                            'pbi_value': '[MISSING]',
                                            'length_mismatch': True,
                                            'column_name': column_name # Get the column name
                                        })
                                elif len(pbi_parts) > len(cognos_parts):
                                    for i in range(min_len, len(pbi_parts)):
                                        column_name = selected_columns[i] if i < len(selected_columns) else f'Segment {i+1}'
                                        mismatch_info.append({
                                            'segment_position': i + 1,
                                            'cognos_value': '[MISSING]',
                                            'pbi_value': pbi_parts[i],
                                            'length_mismatch': True,
                                            'column_name': column_name # Get the column name
                                        })
                                
                                return mismatch_info
                            # Run comparison if both inputs are provided
                            if cognos_key and pbi_key:
                                mismatches = extract_mismatched_segment(cognos_key, pbi_key)

                                if mismatches:
                                    st.write("### ❗ Mismatched Segments Detected:")
                                    for mismatch in mismatches:
                                        st.markdown(f"""
                                        🔸 **Segment #{mismatch['segment_position']}**  
                                        - Column Orignial Name #{mismatch['column_name']}
                                        - Cognos: `{mismatch['cognos_value']}`  
                                        - PBI: `{mismatch['pbi_value']}`  
                                        - {'❗ Length mismatch' if mismatch['length_mismatch'] else '⚠️ Value mismatch'}
                                        """)
                                    st.warning("These differences might be causing validation issues.")
                                else:
                                    st.success("✅ All segments match. No mismatches detected.")    

from typing import List, Optional, Dict

def generate_valid_report(
    pbi_df: pd.DataFrame,
    cognos_df: pd.DataFrame,
    dimension_columns: Optional[List[str]] = None,
) -> pd.DataFrame:
    """
    Generates a validation report comparing two dataframes (PBI and Cognos).

    Args:
        pbi_df: Pandas DataFrame representing PBI data.
        cognos_df: Pandas DataFrame representing Cognos data.
        dimension_columns: Optional list of columns to treat as dimensions.
            If None, all non-numeric columns are used as dimensions.

    Returns:
        Pandas DataFrame representing the validation report.
    """
    try:
        # Handle empty DataFrames
        if pbi_df.empty and cognos_df.empty:
            return pd.DataFrame({"Result": ["Both PBI and Cognos DataFrames are empty"]})
        elif pbi_df.empty:
            return pd.DataFrame({"Result": ["PBI DataFrame is empty"]})
        elif cognos_df.empty:
            return pd.DataFrame({"Result": ["Cognos DataFrame is empty"]})

        # Determine dimension columns if not provided
        if dimension_columns is None:
            numeric_cols_pbi = pbi_df.select_dtypes(include="number").columns
            numeric_cols_cognos = cognos_df.select_dtypes(include="number").columns
            dimension_columns = list(
                pbi_df.columns.difference(numeric_cols_pbi)
            )  # Get non-numeric
            # Ensure the dimension columns exist in both dataframes
            dimension_columns = [
                col
                for col in dimension_columns
                if col in cognos_df.columns
            ]

            if not dimension_columns:  # if there are no common dimension columns.
                return pd.DataFrame(
                    {
                        "Result": [
                            "No common dimension columns found.  Cannot compare."
                        ]
                    }
                )

        # Clean data in dimension columns (handle None, NaN, and empty strings)
        for col in dimension_columns:
            if pbi_df[col].dtype == 'object':
                pbi_df[col] = pbi_df[col].astype(str).str.strip().replace(['None', 'nan', ''], None)
            if cognos_df[col].dtype == 'object':
                cognos_df[col] = cognos_df[col].astype(str).str.strip().replace(['None', 'nan', ''], None)

        # Find differences
        merged_df = pd.merge(
            pbi_df,
            cognos_df,
            how="outer",
            on=dimension_columns,
            suffixes=("_PBI", "_Cognos"),
            indicator=True,
        )

        # Create the validation report
        validation_report = pd.DataFrame()

        for col in dimension_columns:
            validation_report[col] = merged_df[col]

        # Iterate through columns, handling potential errorsRobustly
        for col in pbi_df.columns:
            if col not in dimension_columns:
                cognos_col = col.replace("_PBI", "_Cognos")
                if cognos_col in cognos_df.columns:
                    try:
                        validation_report[f"{col}_PBI"] = merged_df[col]
                        validation_report[f"{cognos_col}_Cognos"] = merged_df[cognos_col]
                        validation_report[f"{col}_Result"] = (
                            merged_df[col] == merged_df[cognos_col]
                        ).map({True: "Pass ✅", False: "Fail ❌"})
                    except KeyError as e:
                        error_message = f"Column '{col}' not found"
                        validation_report[f"{col}_Result"] = f"Error: {error_message}"
                        st.error(f"KeyError: {error_message} - '{col}'.")
                    except Exception as e:
                        error_message = f"Error comparing column '{col}': {type(e).__name__}, {e}"
                        validation_report[f"{col}_Result"] = f"Error: {error_message}"
                        st.error(
                            f"An error of type {type(e).__name__} occurred: {error_message}."
                        )
                else:
                    validation_report[f"{col}_Result"] = "Not Present in Cognos"

        # Add a summary row
        if 'Result' in validation_report:
            total_rows = len(validation_report)
            passed_rows = (validation_report['Result'] == 'Pass ✅').sum()
            validation_report.loc['Total', 'Result'] = total_rows
            validation_report.loc['Passed', 'Result'] = passed_rows
            validation_report.loc['Failed', 'Result'] = total_rows - passed_rows

        return validation_report
    except Exception as e:
        st.error(f"Error in generate_validation_report: {e}")
        return pd.DataFrame()  # Return an empty DataFrame in case of error



def multiplesheets_Validation_Report_Gen():
    # File Upload
    pbi_file = st.file_uploader('Upload PBI Excel file 📉', type='xlsx', key='pbi_upload')
    cognos_file = st.file_uploader('Upload Cognos Excel file 📈', type='xlsx', key='cognos_upload')

    if pbi_file and cognos_file:
        try:
            # Load PBI and Cognos sheets
            pbi_sheets = pd.ExcelFile(pbi_file).sheet_names
            cognos_sheets = pd.ExcelFile(cognos_file).sheet_names

            # Extracting sheets with specific naming pattern
            pbi_pages = [sheet for sheet in pbi_sheets if sheet.startswith('PBI ')]
            cognos_pages = [sheet for sheet in cognos_sheets if sheet.startswith('Cognos ')]

            # Remove prefix and compare
            pbi_suffixes = [sheet.replace('PBI ', '').strip() for sheet in pbi_pages]
            cognos_suffixes = [sheet.replace('Cognos ', '').strip() for sheet in cognos_pages]

            # Matching suffixes
            suffixes = set(pbi_suffixes) & set(cognos_suffixes)

            if not suffixes:
                # Identify unmatched sheet names
                unmatched_pbi = [sheet for sheet in pbi_suffixes if sheet not in cognos_suffixes]
                unmatched_cognos = [sheet for sheet in cognos_suffixes if sheet not in pbi_suffixes]
                if unmatched_pbi:
                    st.warning(f'Unmatched PBI sheets: {unmatched_pbi}')
                if unmatched_cognos:
                    st.warning(f'Unmatched Cognos sheets: {unmatched_cognos}')
                st.warning('No matching suffixes found between PBI and Cognos sheets.')
            else:
                st.success(f'Found {len(suffixes)} matching suffixes.')

                output = io.BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    for suffix in suffixes:
                        pbi_sheet = f'PBI {suffix}'
                        cognos_sheet = f'Cognos {suffix}'

                        try:
                            pbi_df = pd.read_excel(pbi_file, sheet_name=pbi_sheet)
                            cognos_df = pd.read_excel(cognos_file, sheet_name=cognos_sheet)

                            # Get available columns for the current sheet
                            available_columns = list(pbi_df.columns) + list(cognos_df.columns)
                            available_columns = list(set(available_columns))  # Remove duplicates

                            # Add a multi-select for dimension columns, specific to the sheet
                            dimension_columns = st.multiselect(
                                f"Select dimension columns for {suffix}",
                                options=available_columns,
                                default=[],  # Start with no columns selected by default
                            )
                            # Pass dimension_columns to the function
                            validation_report = generate_valid_report(
                                pbi_df, cognos_df, dimension_columns
                            )

                            if not (isinstance(validation_report, tuple) and hasattr(validation_report, 'empty')): #error fix
                                # Save each validation report
                                validation_report.to_excel(writer, sheet_name=f'Validation_Page_{suffix}', index=False)
                            else:
                                st.warning(f"No data to write for sheet: {suffix}")

                        except Exception as e:
                            st.error(f'Error processing sheets with suffix {suffix}: {str(e)}')

                    output.seek(0)
                    today_date = datetime.date.today().strftime('%Y-%m-%d')
                    file_name = f'Validation_Report_{today_date}.xlsx'

                    st.download_button(
                        label='Download Validation Report ⬇️',
                        data=output,
                        file_name=file_name,
                        mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                    )

        except Exception as e:
            st.error(f'Error processing files: {str(e)}')

def analyze_validation_report(file_content, file_type):
    """
    Analyzes a validation report from a CSV or Excel file, handling multiple sheets in Excel.

    Args:
        file_content (bytes): The content of the uploaded file.
        file_type (str): The type of the uploaded file ('csv' or 'xlsx').

    Returns:
        dict: A dictionary containing the analysis results with these keys:
            - 'comparison_df': DataFrame containing PBI and Cognos column names, diffs, Unique ID and Presence.
            - 'error': A string describing any error that occurred, or None if no error.
            - 'presence_analysis': dict containing counts of 'PBI' and 'Cognos' in 'Presence' column, and differing IDs
    """
    
    try:
        if file_type == 'csv':
            # Read the CSV data
            df = pd.read_csv(io.BytesIO(file_content))
        elif file_type == 'xlsx':
            # Read the Excel file, targeting the "Validation_Report" sheet.  Crucially,
            # we do NOT return here if the sheet is not found.  Instead, we let the
            # rest of the analysis proceed, which is what the user wants.  We DO
            # still need to catch the error, so the 'try' block remains.
            try:
                df = pd.read_excel(io.BytesIO(file_content), sheet_name="Validation_Report")
            except KeyError:
                # Instead of returning, set df to an empty DataFrame and continue.
                # This ensures the rest of the code runs. We also set a custom attribute
                # on the DataFrame to indicate that the sheet was not found.
                df = pd.DataFrame()
                df.sheet_not_found = True
        else:
            return {'error': f"Error: Unsupported file type: {file_type}"}

    except Exception as e:
        return {'error': f'Error reading file: {e}'}

    # 1. Identify Matching Prefixes
    pbi_cols = [col for col in df.columns if col.endswith('_PBI')]
    cognos_cols = [col for col in df.columns if col.endswith('_Cognos')]
    diff_cols = [col for col in df.columns if col.endswith('_Diff')]
    id_col = 'unique_key'
    presence_col = 'presence'

    # 2. Extract Common Prefixes
    pbi_prefixes = [col.replace('_PBI', '') for col in pbi_cols]
    cognos_prefixes = [col.replace('_Cognos', '') for col in cognos_cols]
    diff_prefixes = [col.replace('_Diff', '') for col in diff_cols]
    common_prefixes = list(set(pbi_prefixes) & set(cognos_prefixes) & set(diff_prefixes))

    if not common_prefixes:
        return {'error': "Error: No common column prefixes found in the data."}

    # 3. Create Comparison Dataframe
    comparison_df = pd.DataFrame()
    comparison_df['Unique ID'] = df[id_col]
    comparison_df['Presence'] = df[presence_col] 

    for prefix in common_prefixes:
        comparison_df[f'{prefix}_PBI'] = df[f'{prefix}_PBI']
        comparison_df[f'{prefix}_Cognos'] = df[f'{prefix}_Cognos']
        comparison_df[f'{prefix}_Diff'] = df[f'{prefix}_Diff']

    # 4. Analyze 'Presence' Column
    presence_analysis = {}
    if presence_col in df:
        presence_counts = df[presence_col].value_counts()
        presence_analysis['pbi_present_count'] = presence_counts.get('Present in PBI', 0)
        presence_analysis['cognos_present_count'] = presence_counts.get('Present in Cognos', 0)

        # Find Unique IDs where Presence is 'Present in PBI' or 'Present in Cognos'
        pbi_present_ids = df[df[presence_col] == 'Present in PBI'][id_col].tolist()
        cognos_present_ids = df[df[presence_col] == 'Present in Cognos'][id_col].tolist()
        presence_analysis['pbi_present_ids'] = sorted(pbi_present_ids)
        presence_analysis['cognos_present_ids'] = sorted(cognos_present_ids)
    else:
        presence_analysis['pbi_present_count'] = 0
        presence_analysis['cognos_present_count'] = 0
        presence_analysis['pbi_present_ids'] = []
        presence_analysis['cognos_present_ids'] = []

    return {
        'comparison_df': comparison_df,
        'error': None,
        'presence_analysis': presence_analysis,
        'sheet_not_found': getattr(df, 'sheet_not_found', False) # Get custom attribute, default to False
    }


# Main functions for Validation Report Generator
# Function to strip leading zeros and convert to numeric if applicable
def strip_leading_zeros(val):
    try:
        if isinstance(val, str):
            val = val.strip().replace(',', '')
            if val.replace('.', '', 1).isdigit() or \
               (val.startswith('-') and val[1:].replace('.', '', 1).isdigit()):
                return float(val)
        return val
    except:
        return val

# Apply numeric cleaning only on likely numeric columns
def convert_possible_numeric(df):
    for col in df.columns:
        df[col] = df[col].apply(strip_leading_zeros)
        # Try numeric conversion first
        try:
            df[col] = pd.to_numeric(df[col], errors='ignore')
        except:
            pass

        # Then try datetime conversion if it's still an object type
        if df[col].dtype == 'object':
            try:
                # Attempt to parse dates, inferring format
                converted_dates = pd.to_datetime(df[col], errors='coerce')
                # Check if a significant portion of the column was converted to datetime
                non_na_count = df[col].notna().sum()
                successful_conversion_count = converted_dates.notna().sum()
                if successful_conversion_count > 0.8 * non_na_count:  # Convert if at least 80% are likely dates
                    df[col] = converted_dates
            except:
                pass
    return df

# Define the checklist data as a DataFrame
checklist_data = {
    "S.No": range(1, 18),
    "Checklist": [
        "Database & Warehouse is parameterized (In case of DESQL Reports)",
        "All the columns of Cognos replicated in PBI (No extra columns)",
        "All the filters of Cognos replicated in PBI",
        "Filters working as expected (single/multi select as usual)",
        "Column names matching with Cognos",
        "Currency symbols to be replicated",
        "Filters need to be aligned vertically/horizontally",
        "Report Name & Package name to be written",
        "Entire model to be refreshed before publishing to PBI service",
        "Date Last refreshed to be removed from filter/table",
        "Table's column header to be bold",
        "Table style to not have grey bars",
        "Pre-applied filters while generating validation report?",
        "Dateformat to beYYYY-MM-DD [hh:mm:ss] in refresh date as well",
        "Sorting is replicated",
        "Filter pane to be hidden before publishing to PBI service",
        "Mentioned the exception in our validation document like numbers/columns/values mismatch (if any)"
    ],
    "Status - Level1": ["" for _ in range(17)],
    "Status - Level2": ["" for _ in range(17)]
}
checklist_df = pd.DataFrame(checklist_data)

def generate_validation_report(cognos_df, pbi_df, dimension_columns):
    dims = dimension_columns

    cognos_df[dims] = cognos_df[dims].fillna('NAN')
    pbi_df[dims] = pbi_df[dims].fillna('NAN')

    cognos_measures = [col for col in cognos_df.columns if col not in dims and np.issubdtype(cognos_df[col].dtype, np.number)]
    pbi_measures = [col for col in pbi_df.columns if col not in dims and np.issubdtype(pbi_df[col].dtype, np.number)]
    all_measures = list(set(cognos_measures) & set(pbi_measures))

    cognos_agg = cognos_df.groupby(dims)[all_measures].sum().reset_index()
    pbi_agg = pbi_df.groupby(dims)[all_measures].sum().reset_index()

    cognos_agg['unique_key'] = cognos_agg[dims].astype(str).agg('-'.join, axis=1).str.upper()
    pbi_agg['unique_key'] = pbi_agg[dims].astype(str).agg('-'.join, axis=1).str.upper()

    cognos_agg = cognos_agg[['unique_key'] + [col for col in cognos_agg.columns if col != 'unique_key']]
    pbi_agg = pbi_agg[['unique_key'] + [col for col in pbi_agg.columns if col != 'unique_key']]

    validation_report = pd.DataFrame({'unique_key': list(set(cognos_agg['unique_key']) | set(pbi_agg['unique_key']))})

    for dim in dims:
        validation_report[dim] = validation_report['unique_key'].map(dict(zip(cognos_agg['unique_key'], cognos_agg[dim])))
        validation_report[dim].fillna(validation_report['unique_key'].map(dict(zip(pbi_agg['unique_key'], pbi_agg[dim]))), inplace=True)

    validation_report['presence'] = validation_report['unique_key'].apply(
        lambda key: 'Present in Both' if key in cognos_agg['unique_key'].values and key in pbi_agg['unique_key'].values
        else ('Present in Cognos' if key in cognos_agg['unique_key'].values
              else 'Present in PBI')
    )

    for measure in all_measures:
        validation_report[f'{measure}_Cognos'] = validation_report['unique_key'].map(dict(zip(cognos_agg['unique_key'], cognos_agg[measure])))
        validation_report[f'{measure}_PBI'] = validation_report['unique_key'].map(dict(zip(pbi_agg['unique_key'], pbi_agg[measure])))

        validation_report[f'{measure}_Diff'] = validation_report[f'{measure}_PBI'].fillna(0) - validation_report[f'{measure}_Cognos'].fillna(0)

    column_order = ['unique_key'] + dims + ['presence'] + [col for measure in all_measures for col in
                                                            [f'{measure}_Cognos', f'{measure}_PBI', f'{measure}_Diff']]
    validation_report = validation_report[column_order]

    return validation_report, cognos_agg, pbi_agg

def column_checklist(cognos_df, pbi_df):
    cognos_columns = cognos_df.columns.tolist()
    pbi_columns = pbi_df.columns.tolist()

    checklist_df = pd.DataFrame({
        'Cognos Columns': cognos_columns + [''] * (max(len(pbi_columns), len(cognos_columns)) - len(cognos_columns)),
        'PowerBI Columns': pbi_columns + [''] * (max(len(pbi_columns), len(cognos_columns)) - len(pbi_columns))
    })

    checklist_df['Match'] = checklist_df.apply(lambda row: row['Cognos Columns'] == row['PowerBI Columns'], axis=1)

    return checklist_df

def generate_diff_checker(validation_report):
    diff_columns = [col for col in validation_report.columns if col.endswith('_Diff')]

    diff_checker = pd.DataFrame({
        'Diff Column Name': diff_columns,
        'Sum of Difference': [validation_report[col].sum() for col in diff_columns]
    })

    presence_summary = {
        'Diff Column Name': 'All rows present in both',
        'Sum of Difference': 'Yes' if all(validation_report['presence'] == 'Present in Both') else 'No'
    }
    diff_checker = pd.concat([diff_checker, pd.DataFrame([presence_summary])], ignore_index=True)

    return diff_checker

# Dry Run functions
def load_file(uploaded_file):
    if uploaded_file.name.endswith(".csv"):
        return pd.read_csv(uploaded_file)
    else:
        return pd.read_excel(uploaded_file)

def dry_run(file1, file2):
    """
    Performs a dry run comparison between two uploaded files.  It automatically
    identifies and compares numeric columns.

    Args:
        file1 (UploadedFile): The first file to compare.
        file2 (UploadedFile): The second file to compare.
    """
    if file1 is not None and file2 is not None:
        try:
            df1 = load_file(file1)
            df2 = load_file(file2)

            st.subheader("📏 Row Counts")
            col1, col2 = st.columns(2)
            col1.metric("cognos_file Rows", len(df1))
            col2.metric("pbi_file Rows", len(df2))

            # Show common columns
            st.subheader("🧾 Common Columns for Comparison")
            common_cols = list(set(df1.columns) & set(df2.columns))
            st.write(common_cols)

            # **Automatic Column Selection:**
            # Identify common columns that are numeric in both DataFrames
            numeric_cols1 = df1.select_dtypes(include='number').columns
            numeric_cols2 = df2.select_dtypes(include='number').columns
            selected_columns = list(set(numeric_cols1) & set(numeric_cols2))
            st.write(f"Automatically selected numeric columns: {selected_columns}") #show selected columns

            if selected_columns:
                df1_numeric = df1[selected_columns].copy()  # Create a copy to avoid modifying the original DataFrame
                df2_numeric = df2[selected_columns].copy()  # Create a copy
                
                # Convert to numeric, errors='coerce' will turn non-numeric to NaN
                for col in selected_columns:
                    df1_numeric[col] = pd.to_numeric(df1_numeric[col], errors='coerce')
                    df2_numeric[col] = pd.to_numeric(df2_numeric[col], errors='coerce')

                sum1 = df1_numeric.sum(skipna=True)  # Use skipna=True to ignore NaN values
                sum2 = df2_numeric.sum(skipna=True)
                

                percentage_diff = (abs(sum2 - sum1) / sum1) * 100

                comparison = pd.DataFrame({
                    "File 1 Sum": sum1,
                    "File 2 Sum": sum2,
                    "Absolute Difference": (sum1 - sum2).abs(),
                    "% Difference (w.r.t File 1)": percentage_diff.round(2)
                })

                st.subheader("📊 Column-wise Sum Comparison")
                st.dataframe(comparison)

            else:
                st.info("No common numeric columns found to compare.")

        except Exception as e:
            st.error(f"Error: {e}")

    elif file1 is not None or file2 is not None:
        st.warning("Please upload both files to compare.")



def get_db_connection():
    return mysql.connector.connect(
        host="localhost",
        user="root",
        password="shreyshukla12345",
        database="anil_ps"
    )
def fetch_posts_by_owner(owner_email, token):
    headers = {"Authorization": f"Bearer {token}"}
    try:
        res = requests.get(f"{API_URL}/posts-by-owner", params={"email": owner_email}, headers=headers)
        res.raise_for_status()
        return res.json()
    except requests.exceptions.RequestException as e:
        st.error(f"Error fetching posts: {e}")
        return None
    except json.JSONDecodeError as e:
        st.error(f"Error decoding JSON: {e}. Response text: {res.text}")
        return None

def run_child_model_calculation(email, model, input_data, token):
    headers = {"Authorization": f"Bearer {token}"}
    payload = {
        "email": email,
        "model": model,
        "input": input_data
    }
    try:
        res = requests.post(f"{API_URL}/calculate-child-model", json=payload, headers=headers)
        res.raise_for_status()
        return res.json()
    except requests.exceptions.RequestException as e:
        st.error(f"Error during calculation: {e}")
        return None
    except json.JSONDecodeError as e:
        st.error(f"Error decoding JSON: {e}. Response text: {res.text}")
        return None

def fetch_public_posts():
    try:
        res = requests.get(f"{API_URL}/public-posts")
        res.raise_for_status()
        return res.json()
    except requests.exceptions.RequestException as e:
        st.error(f"Error fetching public posts: {e}")
        return None
    except json.JSONDecodeError as e:
        st.error(f"Error decoding JSON: {e}. Response text: {res.text}")
        return None

def fetch_my_posts(token):
    headers = {"Authorization": f"Bearer {token}"}
    try:
        res = requests.get(f"{API_URL}/my-posts", headers=headers)
        res.raise_for_status()
        return res.json()
    except requests.exceptions.RequestException as e:
        st.error(f"Error fetching your posts: {e}")
        return None
    except json.JSONDecodeError as e:
        st.error(f"Error decoding JSON: {e}. Response text: {res.text}")
        return None

# --- UI Sections ---
def login_section():
    """Handles login and signup."""
    if st.session_state.show_signup:
        st.subheader("✨ Sign Up")
        new_email = st.text_input("New Email (must be @goodyear)")
        new_password = st.text_input("New Password", type="password")
        confirm_password = st.text_input("Confirm Password", type="password")

        if st.button("Sign Up"):
            if not new_email.endswith("@goodyear"):
                st.error("Email must end with @goodyear")
            elif new_password != confirm_password:
                st.error("Passwords do not match.")
            elif len(new_password) < 10:
                st.error("Password must be at least 10 characters long")
            elif not new_password.isalnum():
                st.error("Password must not contain special characters")
            else:
                try:
                    res = requests.post(f"{API_URL}/signup", json={"email": new_email, "password": new_password})
                    res.raise_for_status()
                    if res.status_code == 201:
                        st.success("Account created successfully! Please log in.")
                        st.session_state.show_signup = False
                except requests.exceptions.HTTPError as e:
                    error_message = res.json().get("msg", "An error occurred during sign up.")
                    st.error(f"Sign up failed: {error_message}. Status Code: {res.status_code}")
                except requests.exceptions.RequestException as e:
                    st.error(f"Sign up failed. Error: {e}")
                except json.JSONDecodeError as e:
                    st.error(f"Error decoding JSON: {e}. Response text: {res.text}")

        if st.button("Already have an account? Log in"):
            st.session_state.show_signup = False

    else:
        st.subheader("🔐 Login")
        email = st.text_input("Email")
        password = st.text_input("Password", type="password")

        if st.button("Login"):
            try:
                res = requests.post(f"{API_URL}/login", json={"email": email, "password": password})
                res.raise_for_status()
                data = res.json()
                if "access_token" in data:
                    st.session_state.token = data["access_token"]
                    st.session_state.user_email = email  # Store the user email in session state
                    st.session_state.user_id = data.get("user_id")
                    st.success("Logged in successfully!")
                else:
                    st.error("Login failed: 'access_token' not found in response.")
                    st.write("Response received:", data)
            except requests.exceptions.HTTPError as e:
                st.error("Invalid credentials.")
            except requests.exceptions.RequestException as e:
                st.error(f"Login failed. Error: {e}")
            except json.JSONDecodeError as e:
                st.error(f"Error decoding JSON: {e}. Response text: {res.text}")

        if st.button("Don't have an account? Sign up"):
            st.session_state.show_signup = True

def child_model_calculation_section():
# Custom Styling for Better UI
    st.markdown("""
        <style>
            .main-title {
                text-align: center;
                font-size: 28px;
                color: #4CAF50;
                font-weight: bold;
            }
            .sub-title {
                font-size: 22px;
                color: #333333;
            }
            .highlight {
                background-color: #F5F5F5;
                padding: 10px;
                border-radius: 10px;
                margin-bottom: 10px;
            }
            .warning-box {
                background-color: #FFEB3B;
                padding: 10px;
                border-radius: 10px;
            }
        </style>
    """, unsafe_allow_html=True)

    # Authentication Check
    if not st.session_state.get("token"):
        st.warning("🔒 Please login first.")
    else:
        st.markdown('<div class="main-title">🔍 Child Model Calculation</div>', unsafe_allow_html=True)

        # Owner Email Section
        owner_email = st.text_input("📧 Enter Owner Email", value=st.session_state.get("user_email", ""))

        if owner_email:
            posts = fetch_posts_by_owner(owner_email, st.session_state.token)
            if posts:
                st.success(f"✅ Found {len(posts)} posts by {owner_email}")
                for i, post in enumerate(posts, 1):
                    st.markdown(f"<div class='highlight'><h4>{i}. {post['title']}</h4><b>Child Model:</b> {post['type']}<br><b>Calculation:</b><br>{post['content']}</div>", unsafe_allow_html=True)
            else:
                st.info("🚫 No posts found for this owner.")

        st.markdown("<hr>", unsafe_allow_html=True)

        # Child Model Selection
        st.info("📌 Run a new child model calculation manually below.")
        input_owner = st.text_input("📧 Owner Email", value=st.session_state.get("user_email", ""))

        child_model = st.radio("🛠️ Select Child Model", ["Blocked Report", "Playbook Chatbot", "Update Child Model Calculation", "ChatWPDF"])

        # Blocked Report
        if child_model == "Blocked Report":
            username = st.text_input("👤 Enter Username to Fetch Report:")
            if st.button("📊 Fetch Blocked Report"):
                result = run_child_model_calculation(input_owner, child_model, username, st.session_state.token)
                if result:
                    st.success("✅ Blocked Report Result:")
                    st.write(result.get("result", "No result returned."))

        # Playbook Chatbot
        elif child_model == "Playbook Chatbot":
            st.markdown('<div class="sub-title">🔐 Gemini API Key Required</div>', unsafe_allow_html=True)
            user_api_key = st.text_input("🔑 Please enter your Gemini API Key:", type="password")

            if user_api_key:
                try:
                    genai.configure(api_key=user_api_key)
                    embedding_model = genai.embed_content

                    # Upload Excel File
                    uploaded_file = st.file_uploader("📂 Upload your Excel file", type=["xlsx", "xls"])

                    if uploaded_file:
                        df = pd.read_excel(uploaded_file)
                        df.rename(columns={"Source": "owner", "Child model name": "Child_model_name", "Calculations present": "expression", "PBI Column Name": "pbi_column"}, inplace=True)
                        df.dropna(subset=["expression"], inplace=True)

                        st.write("### 🔎 Preview of the Excel Data")
                        st.dataframe(df)

                        # Embed Content
                        def get_embedding(text):
                            response = embedding_model(model="models/embedding-001", content=text, task_type="semantic_similarity")
                            return response["embedding"]

                        df["embedding"] = df["expression"].apply(get_embedding)
                        embedding_matrix = np.vstack(df["embedding"].values)

                        st.title("🧠 DAX Expression Chatbot")
                        user_input = st.text_input("💬 Ask a question or enter a DAX expression:")

                        if user_input:
                            user_embedding = get_embedding(user_input)
                            similarities = cosine_similarity([user_embedding], embedding_matrix)[0]
                            best_idx = np.argmax(similarities)

                            result = {
                                "Matched Expression": df.iloc[best_idx]["expression"],
                                "Owner": df.iloc[best_idx]["owner"],
                                "PBI Column": df.iloc[best_idx]["pbi_column"],
                                "Similarity Score": round(similarities[best_idx], 4)
                            }

                            st.markdown(f"<div class='highlight'><b>Matched Expression:</b> {result['Matched Expression']}<br><b>Owner:</b> {result['Owner']}<br><b>PBI Column:</b> {result['PBI Column']}<br><b>Similarity Score:</b> {result['Similarity Score']}</div>", unsafe_allow_html=True)

                except Exception as e:
                    st.error(f"❌ Failed to configure Gemini API. Error: {e}")
            else:
                st.warning("⚠️ Please enter your Gemini API key above to continue.")

        # Update Child Model Calculation
        elif child_model == "Update Child Model Calculation":
            st.markdown('<div class="sub-title">📄 Update Excel Sheet - "Calculations present" Column</div>', unsafe_allow_html=True)

            df = pd.read_excel("DummyExcel.xlsx")
            df.rename(columns={"Source": "owner", "Child model name": "Child_model_name", "Calculations present": "expression", "PBI Column Name": "pbi_column"}, inplace=True)
            st.write("📄 Current Entries:", df.head())

            owner_input = st.text_input("🔍 Enter Owner (Source column)")
            model_name_input = st.text_input("📊 Enter Child Model Name")
            new_calc = st.text_area("✍️ Enter New Calculation Expression")

            if st.button("🔄 Update Calculation"):
                match_idx = df[(df["owner"] == owner_input) & (df["Child_model_name"] == model_name_input)].index

                if not match_idx.empty:
                    df.at[match_idx[0], "expression"] = new_calc
                    try:
                        book = load_workbook("Book1.xlsx")
                        with pd.ExcelWriter("Book1.xlsx", engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
                            writer.book = book
                            df.to_excel(writer, index=False, sheet_name="Sheet1")
                        st.success("✅ Excel updated successfully!")
                    except Exception as e:
                        st.error(f"❌ Error writing to Excel: {e}")
                else:
                    st.warning("⚠️ No matching entry found. Please check Owner and Child Model Name.")

        # ChatWPDF
        elif child_model == "ChatWPDF":
            st.title("📄 Chat with Your PDF Using Gemini (No Vector DB)")

            api_key_input = st.text_input("🔑 Enter your Gemini API key:", type="password")
            uploaded_files = st.file_uploader("📂 Upload PDF files", type="pdf", accept_multiple_files=True)

            if api_key_input:
                genai.configure(api_key=api_key_input)

            if uploaded_files:
                context = "\n".join([PdfReader(file).pages[0].extract_text() for file in uploaded_files])
                st.success("✅ PDF text extracted!")

                gemini = genai.GenerativeModel().start_chat(history=[])
                user_question = st.text_input("💬 Ask a question based on the uploaded PDF")

                if user_question:
                    response = gemini.send_message(f"Using this context: {context}\n\nQuestion: {user_question}").text
                    st.markdown("### 📝 Answer:")
                    st.write(response)
            else:
                st.info("🔍 Please upload PDF files to start.")
    # def authenticate_user(username, password):
    #     db_connection = get_db_connection()
    #     if db_connection is not None and db_connection.is_connected():
    #         try:
    #             cursor = db_connection.cursor()
    #             sql = "SELECT id FROM users WHERE username = %s AND password_hash = %s"
    #             cursor.execute(sql, (username, password))  # Ensure password is securely hashed in a real app
    #             result = cursor.fetchone()
    #             cursor.close()
                
    #             if result:
    #                 user_id = result[0]
    #                 st.session_state.user_id = user_id  # Store in session state
    #                 return True
    #             else:
    #                 return False
    #         except mysql.connector.Error as e:
    #             st.sidebar.error(f"Database error: {e}")
    #     return False

def findings_section():
    st.subheader("🔍 Browse Findings")
    view_option = st.radio("Choose what to do:", ["Issues To Public", "Issues Private Findings", "Add Finding related to resolved Issue"])

    if view_option == "Issues To Public(Was Unable to solve..)":
        posts = fetch_public_posts()
        if posts:
            for post in posts:
                if post["type"] == "Findings":
                    with st.expander(f"{post['title']} ({post['type']})"):
                        st.write(post["content"])
                        st.caption(f"Author: {post['username']} | Visibility: Public")

    elif view_option == "Issues Private Findings":
        if not st.session_state.token:
            st.warning("Please login to view your private findings.")
        else:
            posts = fetch_my_posts(st.session_state.token)
            if posts:
                private_findings = [p for p in posts if not p["is_public"] and p["type"] == "Findings"]
                if not private_findings:
                    st.info("You don't have any private findings.")
                else:
                    for post in private_findings:
                        with st.expander(f"{post['title']} ({post['type']})"):
                            st.write(post["content"])
                            st.caption("Visibility: Private")

    elif view_option == "Add Finding related to resolved Issue":
        st.markdown("""
            <style>
                .main-title {
                    text-align: center;
                    font-size: 28px;
                    color: #4CAF50;
                    font-weight: bold;
                }
                .sub-title {
                    font-size: 22px;
                    color: #333333;
                }
                .highlight {
                    background-color: #F5F5F5;
                    padding: 10px;
                    border-radius: 10px;
                    margin-bottom: 10px;
                }
                .warning-box {
                    background-color: #FFEB3B;
                    padding: 10px;
                    border-radius: 10px;
                }
            </style>
        """, unsafe_allow_html=True)

        # Authentication Check
        if not st.session_state.get("token"):
            st.warning("🔒 Please login to add a finding.")
            st.stop()

        if "user_email" not in st.session_state:
            st.warning("⚠️ Email not found. Please log in first.")
            st.stop()

        st.markdown('<div class="main-title">📝 Create a New Finding Post</div>', unsafe_allow_html=True)

        # User Inputs for the Finding Post
        with st.form("finding_form"):
            user_id = st.text_input("📧 Your Email:", value=st.session_state.get("user_email", ""))
            title = st.text_input("📝 Title of the Finding:")
            content = st.text_area("🔍 Detailed Content:")
            severity = st.selectbox("⚠️ Severity Level", ["Low", "Medium", "High", "Critical"])
            impact = st.text_area("💥 Describe the Impact:")
            is_public = st.checkbox("🌍 Make this finding public?", value=False)
            
            submit_button = st.form_submit_button("🚀 Submit Finding")

        if submit_button:
            if not title.strip() or not content.strip() or not impact.strip():
                st.warning("⚠️ Please fill all fields before submitting.")
            else:
                try:
                    conn = get_db_connection()
                    cursor = conn.cursor()

                    # Step 1: Insert into posts table
                    cursor.execute("""
                        INSERT INTO posts (title, content, type, is_public, owner_email, user_id)
                        VALUES (%s, %s, %s, %s, %s, %s)
                    """, (title, content, "Findings", is_public, user_id, user_id))

                    post_id = cursor.lastrowid  # Get the inserted post's ID

                    # Step 2: Ensure findings table exists
                    cursor.execute("""
                        CREATE TABLE IF NOT EXISTS findings (
                            id INT AUTO_INCREMENT PRIMARY KEY,
                            post_id INT UNIQUE NOT NULL,
                            severity VARCHAR(50),
                            impact TEXT,
                            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                            FOREIGN KEY (post_id) REFERENCES posts(id)
                        );
                    """)

                    # Step 3: Insert into findings table
                    cursor.execute("""
                        INSERT INTO findings (post_id, severity, impact)
                        VALUES (%s, %s, %s)
                    """, (post_id, severity, impact))

                    conn.commit()
                    cursor.close()
                    conn.close()

                    st.success("✅ Finding submitted successfully!")
                except mysql.connector.Error as e:
                    st.error(f"❌ Error occurred while submitting the finding: {e}")
def logout_section():
    st.session_state.token = None
    st.session_state.show_signup = False
    st.session_state.user_email = None  # Clear user email on logout
    st.session_state.user_id = None
    st.success("Logged out successfully.")

def main():
    """Main function to run the Streamlit app."""
    st.markdown("""
    <style>
        /* Global Background */
        body {
            background-color: #121212; /* Dark background */
            color: #d1f7c4; /* Light green text */
        }

        /* Sidebar Styling */
        .sidebar .sidebar-content {
            background-color: #1f2a40;
            color: #d1f7c4;
        }

        .sidebar .sidebar-content .element-container {
            padding: 10px 15px;
            margin-bottom: 5px;
            border-radius: 8px;
            transition: background-color 0.3s ease;
        }

        .sidebar .sidebar-content .element-container:hover {
            background-color: #3a4b6c;
        }

        /* History Panel */
        .history-panel {
            background-color: #3a4b6c;
            border-radius: 8px;
            padding: 10px;
        }

        .history-entry {
            padding: 5px 0;
            border-bottom: 1px solid #d1f7c4;
        }

        .history-entry:last-child {
            border-bottom: none;
        }

        /* Highlight */
        .highlight {
            background-color: #3a4b6c;
            border-radius: 5px;
            padding: 5px;
            margin-bottom: 5px;
        }
    </style>
    """, unsafe_allow_html=True)

    st.session_state.setdefault("page_history", [])
    st.session_state.setdefault("main_navigation_selectbox", "Login")
    st.session_state.setdefault("logged_in", False)
    st.session_state.setdefault("user_id", None)

    # Menu options with icons
    menu_options = {
        "🏠 Login": "Login",
        "📊 Child Model Calculation": "Child Model Calculation",
        "🛠️ Error Findings": "Error Findings",
        "📄 Generate Validation Report": "Generate Validation Report",
        "🗂️ Multiple Sheets Validation": "multiplesheets_Validation_Report_Gen",
        "🧬 Excel Combiner": "Excel Combiner",
        "🌐 Web Summarization": "YT and WebSummarization through Link(model small so not wrking)",
        "🔧 Parameter Generator": "param_gen",
        "🚪 Logout": "Logout"
    }

    # Sidebar navigation
    with st.sidebar:
        st.header("Navigation")
        menu_selection = st.selectbox(
            "Navigate",
            list(menu_options.keys()),
            index=list(menu_options.values()).index(st.session_state.main_navigation_selectbox)
        )
        st.session_state.main_navigation_selectbox = menu_options[menu_selection]

        # Display recent visits
        st.subheader("Recently Visited")
        with st.expander("History"):
            if st.session_state.page_history:
                st.caption("Click to navigate to previously visited pages")
                for i, entry in enumerate(reversed(st.session_state.page_history)):
                    is_current = entry["page"] == st.session_state.main_navigation_selectbox
                    button_label = f"✅ {entry['page']}" if is_current else entry["page"]
                    if st.button(button_label, key=f"history_{i}_{entry['page']}", disabled=is_current):
                        st.session_state.main_navigation_selectbox = entry["page"]
                        st.rerun()
                    if i < len(st.session_state.page_history) - 1:
                        st.markdown("---")
            else:
                st.info("No history available.")

    # Page visit logging logic
    now = datetime.datetime.now()
    timestamp_str = now.strftime("%Y-%m-%d %H:%M:%S")
    current_visit = {"page": st.session_state.main_navigation_selectbox, "time": timestamp_str}

    if not st.session_state.page_history or st.session_state.page_history[-1]["page"] != st.session_state.main_navigation_selectbox:
        st.session_state.page_history.append(current_visit)

    # Limit history to last 10 visits
    st.session_state.page_history = st.session_state.page_history[-10:]

    # Log visit to database
    user_id = st.session_state.get("user_id")
    db_connection = get_db_connection()

    if user_id and db_connection is not None and db_connection.is_connected():
        try:
            cursor = db_connection.cursor()
            sql = "INSERT INTO page_visits (page_name, visit_time, user_id) VALUES (%s, %s, %s)"
            cursor.execute(sql, (current_visit["page"], now, user_id))
            db_connection.commit()
            cursor.close()
        except mysql.connector.Error as e:
            st.sidebar.error(f"Error saving visit to DB: {e}")
        finally:
            db_connection.close()

    # Page selection logic
    selected_page = st.session_state.main_navigation_selectbox
    if selected_page == "Login":
        login_section()
    elif selected_page == "Child Model Calculation":
        child_model_calculation_section()
    elif selected_page == "Error Findings":
        findings_section()
    elif selected_page == "param_gen":
        param_gen()
    elif selected_page == "multiplesheets_Validation_Report_Gen":
        multiplesheets_Validation_Report_Gen()
    elif selected_page == "Generate Validation Report":
        Generate_Validation_Report()
    elif selected_page == "Excel Combiner":
        Excel_Combiner()
    elif selected_page == "YT and WebSummarization through Link(model small so not wrking)":
        YT_and_WebSummarization_through_Link()
    elif selected_page == "Logout":
        logout_section()

if __name__ == "__main__":
    main()
#     st.markdown("""
#         <style>
#             .sidebar .sidebar-content {
#                 background-color: #1f2a40;
#                 color: #ffffff;
#             }
#             .sidebar .sidebar-content .element-container {
#                 padding: 10px 15px;
#                 margin-bottom: 5px;
#                 border-radius: 8px;
#                 transition: background-color 0.3s ease;
#             }
#             .sidebar .sidebar-content .element-container:hover {
#                 background-color: #3a4b6c;
#             }
#             .history-panel {
#                 background-color: #3a4b6c;
#                 border-radius: 8px;
#                 padding: 10px;
#             }
#             .history-entry {
#                 padding: 5px 0;
#                 border-bottom: 1px solid #ffffff;
#             }
#             .history-entry:last-child {
#                 border-bottom: none;
#             }
#             .highlight {
#                 background-color: #607D8B;
#                 border-radius: 5px;
#                 padding: 5px;
#                 margin-bottom: 5px;
#             }
#         </style>
#     """)

#     # Initialize session state variables
#     if 'page_history' not in st.session_state:
#         st.session_state.page_history = []

#     if 'main_navigation_selectbox' not in st.session_state:
#         st.session_state.main_navigation_selectbox = "Login"

#     if 'logged_in' not in st.session_state:
#         st.session_state.logged_in = False

#     if 'user_id' not in st.session_state:
#         st.session_state.user_id = None

#     # Menu options with icons
#     menu_options = [
#         "🏠 Login",
#         "📊 Child Model Calculation",
#         "🛠️ Error Findings",
#         "📄 Generate Validation Report",
#         "🗂️ Multiple Sheets Validation",
#         "🧬 Excel Combiner",
#         "🌐 Web Summarization",
#         "🔧 Parameter Generator",
#         "🚪 Logout"
#     ]

#     # Sidebar navigation
#     with st.sidebar:
#         st.header("Navigation")
#         menu_selection = st.selectbox(
#             "Navigate",
#             menu_options,
#             index=menu_options.index(st.session_state.main_navigation_selectbox)
#         )

#         st.session_state.main_navigation_selectbox = menu_selection

#         # Display recent visits
#         st.subheader("Recently Visited")
#         with st.expander("History"):
#             if st.session_state.page_history:
#                 for entry in reversed(st.session_state.page_history):
#                     is_current = entry['page'] == menu_selection
#                     button_label = f"✅ {entry['page']}" if is_current else entry['page']
#                     st.button(button_label, key=f"history_{entry['page']}", disabled=is_current)
#             else:
#                 st.info("No history available.")

#     # Page visit logic
#     now = datetime.datetime.now()
#     timestamp_str = now.strftime("%Y-%m-%d %H:%M:%S")
#     current_visit = {"page": menu_selection, "time": timestamp_str}

#     if not st.session_state.page_history or st.session_state.page_history[-1]['page'] != menu_selection:
#         st.session_state.page_history.append(current_visit)

#     st.session_state.page_history = st.session_state.page_history[-10:]

#     # Log visit to DB
#     db_connection = get_db_connection()
#     user_id = st.session_state.get("user_id", None)

#     if db_connection is not None and db_connection.is_connected():
#         try:
#             cursor = db_connection.cursor()
#             cursor.execute(
#                 "INSERT INTO page_visits (page_name, visit_time, user_id) VALUES (%s, %s, %s)",
#                 (current_visit['page'], now, user_id)
#             )
#             db_connection.commit()
#             cursor.close()
#         except mysql.connector.Error as e:
#             st.sidebar.error(f"Error saving visit: {e}")
#         finally:
#             db_connection.close()
#     else:
#         st.sidebar.warning("Not connected to DB. Visit not saved.")

#     # """Main function to run the Streamlit app."""

#     # # Initialize session state variables
#     # if 'page_history' not in st.session_state:
#     #     st.session_state.page_history = []

#     # if 'main_navigation_selectbox' not in st.session_state:
#     #     st.session_state.main_navigation_selectbox = "Login"  # Default starting page

#     # if 'logged_in' not in st.session_state:
#     #     st.session_state.logged_in = False  # Default login state

#     # if 'user_id' not in st.session_state:
#     #     st.session_state.user_id = None  # Ensure user_id is initialized

#     # # Debugging message
#     # # st.write("Session state before navigation:", st.session_state)

#     # # Define menu options
#     # menu_options = [
#     #     "Login",
#     #     "Child Model Calculation",
#     #     "Error Findings",
#     #     "Generate Validation Report",
#     #     "multiplesheets_Validation_Report_Gen",
#     #     "Excel Combiner",
#     #     "YT and WebSummarization through Link(model small so not wrking)",
#     #     "param_gen",
#     #     "Logout"
#     # ]

#     # # Sidebar navigation
    
#     # menu_selection = st.sidebar.selectbox(
#     #     "Navigate",
#     #     menu_options,
#     #     index=menu_options.index(st.session_state.main_navigation_selectbox)
#     # )

#     # # Update session state
#     # st.session_state.main_navigation_selectbox = menu_selection

#     # # Get current timestamp
#     # now = datetime.datetime.now()
#     # timestamp_str = now.strftime("%Y-%m-%d %H:%M:%S")

#     # # Create a history entry
#     # current_visit = {"page": menu_selection, "time": timestamp_str}

#     # # Update page history
#     # if not st.session_state.page_history or st.session_state.page_history[-1]['page'] != menu_selection:
#     #     st.session_state.page_history.append(current_visit)

#     # # Limit history length
#     # st.session_state.page_history = st.session_state.page_history[-10:]

#     # # Display history in sidebar
#     # with st.sidebar:
#     #     with st.expander("Recently Visited"):
#     #         if st.session_state.page_history:
#     #             st.caption("Click to navigate to previously visited pages")
#     #             for i, entry in enumerate(reversed(st.session_state.page_history)):
#     #                 st.markdown(f"`{entry['time']}`")
#     #                 is_current_page = entry['page'] == menu_selection
#     #                 button_label = f"📍 {entry['page']} (current)" if is_current_page else entry['page']
#     #                 if st.button(
#     #                     button_label, key=f"history_btn_{i}_{entry['page']}",
#     #                     help=f"Go to {entry['page']} visited at {entry['time']}",
#     #                     disabled=is_current_page
#     #                 ):
#     #                     st.session_state.main_navigation_selectbox = entry['page']
#     #                     st.rerun()
#     #                 if i < len(st.session_state.page_history) - 1:
#     #                     st.markdown("---")
#     #         else:
#     #             st.info("No history yet.")

#     # # Log visit to database
#     # db_connection = get_db_connection()
#     # user_id = st.session_state.get("user_id", None)

#     # # if user_id is None:
#     # #     st.error("User ID is missing. Cannot log visit.")
#     # if db_connection is not None and db_connection.is_connected():
#     #     try:
#     #         cursor = db_connection.cursor()
#     #         sql = "INSERT INTO page_visits (page_name, visit_time, user_id) VALUES (%s, %s, %s)"
#     #         cursor.execute(sql, (current_visit['page'], now, user_id))
#     #         db_connection.commit()
#     #         cursor.close()
#     #     except mysql.connector.Error as e:
#     #         st.sidebar.error(f"Error saving visit to DB: {e}")
#     #     finally:
#     #         db_connection.close()
#     # else:
#     #     st.sidebar.warning("Not connected to database. Visit not saved.")

#     # Debugging message
#     # st.write("Session state after navigation:", st.session_state)

#     # Page selection logic
#     if menu_selection == "Login":
#         login_section()
#     elif menu_selection == "Child Model Calculation":
#         child_model_calculation_section()
#     elif menu_selection == "Error Findings":
#         findings_section()
#     elif menu_selection == "param_gen":
#         param_gen()
#     elif menu_selection == "multiplesheets_Validation_Report_Gen":
#         multiplesheets_Validation_Report_Gen()
#     elif menu_selection == "Generate Validation Report":
#         Generate_Validation_Report()
#     elif menu_selection == "Excel Combiner":
#         Excel_Combiner()
#     elif menu_selection == "YT and WebSummarization through Link(model small so not wrking)":
#         YT_and_WebSummarization_through_Link()
#     elif menu_selection == "Logout":
#         logout_section()

# if __name__ == "__main__":
#     main()
